from __future__ import annotations

import os
from datetime import datetime, timezone
from typing import Any, Dict

from flask import Flask, jsonify, request, send_from_directory
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
XLSX_PATH = os.path.join(DATA_DIR, "registrations.xlsx")

FRONTEND_DIR = os.path.abspath(os.path.join(BASE_DIR, ".."))


def ensure_workbook() -> None:
    os.makedirs(DATA_DIR, exist_ok=True)
    if os.path.exists(XLSX_PATH):
        return

    wb = Workbook()
    # Participants sheet
    ws = wb.active
    ws.title = "Participants"
    ws.append(["Timestamp (UTC)", "Full name", "Email", "Affiliation", "Category", "Primary interest", "Experience level", "Notes"])

    # Sponsors sheet
    ws2 = wb.create_sheet("Sponsors")
    ws2.append(["Timestamp (UTC)", "Organization", "Contact person", "Email", "Phone", "Interest", "Message"])

    wb.save(XLSX_PATH)


def autosize(ws) -> None:
    # Simple column autosize based on max cell length
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 52)


def append_row(sheet_name: str, values: list[Any]) -> None:
    ensure_workbook()
    wb = load_workbook(XLSX_PATH)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]
    ws.append(values)
    autosize(ws)
    wb.save(XLSX_PATH)


app = Flask(__name__, static_folder=FRONTEND_DIR, static_url_path="")


@app.get("/")
def root():
    # Serve the website
    return send_from_directory(FRONTEND_DIR, "index.html")


@app.get("/assets/<path:filename>")
def assets_route(filename: str):
    return send_from_directory(os.path.join(FRONTEND_DIR, "assets"), filename)


@app.post("/api/register")
def api_register():
    data: Dict[str, Any] = request.get_json(silent=True) or {}
    required = ["name", "email", "affiliation", "category", "interest", "level"]
    missing = [k for k in required if not str(data.get(k, "")).strip()]
    if missing:
        return jsonify({"ok": False, "error": f"Missing fields: {', '.join(missing)}"}), 400

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    append_row(
        "Participants",
        [
            ts,
            str(data.get("name", "")).strip(),
            str(data.get("email", "")).strip(),
            str(data.get("affiliation", "")).strip(),
            str(data.get("category", "")).strip(),
            str(data.get("interest", "")).strip(),
            str(data.get("level", "")).strip(),
            str(data.get("notes", "")).strip(),
        ],
    )
    return jsonify({"ok": True})


@app.post("/api/sponsor")
def api_sponsor():
    data: Dict[str, Any] = request.get_json(silent=True) or {}
    required = ["org", "contact", "sponsorEmail", "interestLevel"]
    missing = [k for k in required if not str(data.get(k, "")).strip()]
    if missing:
        return jsonify({"ok": False, "error": f"Missing fields: {', '.join(missing)}"}), 400

    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    append_row(
        "Sponsors",
        [
            ts,
            str(data.get("org", "")).strip(),
            str(data.get("contact", "")).strip(),
            str(data.get("sponsorEmail", "")).strip(),
            str(data.get("phone", "")).strip(),
            str(data.get("interestLevel", "")).strip(),
            str(data.get("message", "")).strip(),
        ],
    )
    return jsonify({"ok": True})


@app.get("/api/download/xlsx")
def download_xlsx():
    ensure_workbook()
    return send_from_directory(DATA_DIR, "registrations.xlsx", as_attachment=True)


if __name__ == "__main__":
    # Run: python app.py
    ensure_workbook()
    app.run(host="0.0.0.0", port=5000, debug=True)
